"""Asigna boxes a médicos desde la columna BOX de la hoja 'LISTADO DE MEDICOS'
de AGENDA.xlsx. Matchea médicos por RUT (deben existir ya en el sistema, ver
`manage.py reiniciar_medicos`).

Formato de la columna BOX observado en esta hoja (distinto al de 'Médicos CM'):
sin comas, con grupos "día(s)+box" separados por espacios y a veces sin espacio
entre ellos, ej. "LJ6 M7 MIE 4V1" = Lunes y Jueves→Box 6, Martes→Box 7,
Miércoles→Box 4, Viernes→Box 1. Números sueltos como "22" aplican todos los
días. Textos sin número (ej. "Ecografias") se toman como box con ese nombre.

Filas con anotaciones ambiguas ("o", "libre", paréntesis, o un número inicial
sin día que lo preceda) se omiten y se listan al final para asignar a mano —
no se adivina.

Es idempotente (usa get_or_create): se puede volver a correr sin duplicar.

Uso: python manage.py asignar_boxes_listado AGENDA.xlsx [--hoja "LISTADO DE MEDICOS"]
"""

import re
import unicodedata

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from medicos.models import Medico

from ...models import AsignacionBox, Box

DIA_3 = {"LUN": 0, "MAR": 1, "MIE": 2, "JUE": 3, "VIE": 4, "SAB": 5, "DOM": 6}
DIA_2 = {"MA": 1, "MI": 2, "JU": 3, "VI": 4, "SA": 5, "DO": 6, "LU": 0}
DIA_1 = {"L": 0, "M": 1, "J": 3, "V": 4, "S": 5, "D": 6, "X": 2}
NOMBRES_DIA = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

MARCADORES_AMBIGUOS = ("(", ")", " O ", "LIBRE")
PISO1_MAXIMO = 7  # Box 1 a 7 = Piso 1 (trauma). Box 8+ y boxes con nombre = Piso 2.


def _piso_para(nombre_box):
    m = re.match(r"^Box\s+(\d+)$", nombre_box)
    if m and int(m.group(1)) <= PISO1_MAXIMO:
        return 1
    return 2


def _normalizar(texto):
    sin_acentos = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return sin_acentos.upper()


def _decodificar_dias(letras):
    dias = []
    i, n = 0, len(letras)
    while i < n:
        tres, dos, uno = letras[i:i + 3], letras[i:i + 2], letras[i:i + 1]
        if tres in DIA_3:
            dias.append(DIA_3[tres]); i += 3
        elif dos in DIA_2:
            dias.append(DIA_2[dos]); i += 2
        elif uno in DIA_1:
            dias.append(DIA_1[uno]); i += 1
        else:
            return None
    return dias


def parsear_box(valor):
    """Devuelve lista de (box_nombre, dia_semana|None), o None si es ambiguo."""
    if valor is None:
        return []
    if isinstance(valor, (int, float)):
        if isinstance(valor, float) and not valor.is_integer():
            return None
        return [(f"Box {int(valor)}", None)]

    texto = str(valor).strip()
    if not texto:
        return []

    mayus = _normalizar(texto)
    if any(m in f" {mayus} " for m in MARCADORES_AMBIGUOS):
        return None

    # Quita espacios salvo cuando están entre dos dígitos (eso sí es ambiguo,
    # ej. 'Mié27 5' no debe fusionarse en 'Box 275').
    sin_espacios = re.sub(r"(?<!\d)\s+|\s+(?!\d)", "", mayus)

    if sin_espacios.isdigit():
        return [(f"Box {int(sin_espacios)}", None)]

    if not any(c.isdigit() for c in sin_espacios):
        return [(texto.strip().upper(), None)]

    if sin_espacios[0].isdigit():
        return None  # número inicial sin día que lo preceda: ambiguo

    resultado = []
    i, n = 0, len(sin_espacios)
    while i < n:
        li = i
        while i < n and not sin_espacios[i].isdigit():
            i += 1
        letras = sin_espacios[li:i]
        if not letras:
            return None
        di = i
        while i < n and sin_espacios[i].isdigit():
            i += 1
        digitos = sin_espacios[di:i]
        if not digitos:
            return None
        dias = _decodificar_dias(letras)
        if dias is None:
            return None
        box_nombre = f"Box {int(digitos)}"
        for d in dias:
            resultado.append((box_nombre, d))
    return resultado


class Command(BaseCommand):
    help = "Asigna boxes a médicos desde la columna BOX de 'LISTADO DE MEDICOS' (AGENDA.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str)
        parser.add_argument("--hoja", type=str, default="LISTADO DE MEDICOS")
        parser.add_argument("--fila-inicio", type=int, default=2)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        try:
            wb = openpyxl.load_workbook(options["archivo"], data_only=True)
        except (OSError, KeyError) as exc:
            raise CommandError(f"No se pudo abrir '{options['archivo']}': {exc}")

        try:
            ws = wb[options["hoja"]]
        except KeyError:
            raise CommandError(
                f"La hoja '{options['hoja']}' no existe. Hojas disponibles: {wb.sheetnames}"
            )

        asignadas, medicos_sin_match, ambiguos = 0, [], []

        with transaction.atomic():
            for fila in ws.iter_rows(min_row=options["fila_inicio"], values_only=True):
                nombre = (fila[0] or "").strip() if len(fila) > 0 else ""
                if not nombre:
                    continue
                rut = str(fila[1] or "").strip() if len(fila) > 1 else ""
                box_valor = fila[6] if len(fila) > 6 else None

                medico = Medico.objects.filter(rut=rut).first() if rut else None
                if medico is None:
                    if box_valor:
                        medicos_sin_match.append(nombre)
                    continue

                reglas = parsear_box(box_valor)
                if reglas is None:
                    ambiguos.append((nombre, box_valor))
                    continue

                for box_nombre, dia_semana in reglas:
                    box_obj, _ = Box.objects.get_or_create(
                        nombre=box_nombre, defaults={"piso": _piso_para(box_nombre)}
                    )
                    _, creada = AsignacionBox.objects.get_or_create(
                        medico=medico, box=box_obj, dia_semana=dia_semana
                    )
                    asignadas += int(creada)

            if options["dry_run"]:
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(f"Asignaciones de box creadas: {asignadas}."))
        if ambiguos:
            self.stdout.write(self.style.WARNING(
                "Filas con BOX ambiguo (asignar a mano en el admin):"
            ))
            for nombre, valor in ambiguos:
                self.stdout.write(f"  - {nombre}: BOX = {valor!r}")
        if medicos_sin_match:
            self.stdout.write(self.style.WARNING(
                "Médicos con box en la planilla pero sin coincidencia por RUT en el sistema:"
            ))
            for nombre in medicos_sin_match:
                self.stdout.write(f"  - {nombre}")
        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-run: no se guardó nada (rollback)."))
