"""Importa los archivos 'ExportarOferta' y 'ExportarBloqueos' del sistema externo.

Ambos archivos no traen RUT del profesional, solo 'NombreRecurso', y con el nombre en
otro orden que el usado en el mantenedor de médicos (ej. "Pablo Andres Giannini" vs.
"GIANNINI PABLO ANDRES"). El matching con Medico se hace normalizando ambos nombres
(sin tildes, en mayúsculas, comparando el conjunto de palabras sin importar el orden).
Filas cuyo recurso no coincide con ningún médico registrado (equipos, salas, u otros
profesionales aún no cargados) se omiten.

Cada importación reemplaza por completo los datos anteriores: los archivos representan
el estado vigente de la oferta/bloqueos, no un historial acumulable.
"""

import datetime
import io
import unicodedata

import openpyxl
from django.db import transaction

from medicos.models import Medico

from .models import BloqueoMedico, OfertaMedico


def _normalizar_nombre(texto):
    texto = unicodedata.normalize("NFKD", texto.upper()).encode("ascii", "ignore").decode("ascii")
    return frozenset(t for t in texto.split() if t)


def _mapa_medicos_por_nombre():
    return {_normalizar_nombre(m.nombre_completo): m for m in Medico.objects.all()}


def _parsear_hora(texto):
    horas, minutos = texto.split(":")
    return datetime.time(int(horas), int(minutos))


def _cargar_libro(contenido_bytes):
    return openpyxl.load_workbook(io.BytesIO(contenido_bytes), data_only=True)


def _es_consulta(nombre_servicio):
    return _normalizar_comparable(nombre_servicio).startswith("consulta")


def _normalizar_comparable(texto):
    sin_acentos = unicodedata.normalize("NFKD", texto or "").encode("ascii", "ignore").decode("ascii")
    return sin_acentos.strip().lower()


def _ofertas_de_consulta(wb):
    """IDs de oferta (bloque horario) que corresponden a consultas médicas, según la
    hoja ServicioOferta. Excluye bloques exclusivos para imágenes/procedimientos/
    vacunas, que nunca aparecen como cita en la agenda diaria (filtrada por
    Área=Consulta Presencial) y no deben contarse como oferta de citas."""
    ws_servicios = wb["ServicioOferta"]
    con_servicio = {}
    for fila in ws_servicios.iter_rows(min_row=2, values_only=True):
        id_oferta, nombre_servicio = fila[0], fila[3]
        if not id_oferta or not nombre_servicio:
            continue
        con_servicio[id_oferta] = con_servicio.get(id_oferta, False) or _es_consulta(nombre_servicio)
    return con_servicio


def importar_oferta_xlsx(contenido_bytes):
    wb = _cargar_libro(contenido_bytes)
    ws_ofertas = wb["Ofertas"]
    ws_horarios = wb["HorariosOfertas"]
    medicos_por_nombre = _mapa_medicos_por_nombre()
    ofertas_consulta = _ofertas_de_consulta(wb)

    medico_por_oferta = {}
    for fila in ws_ofertas.iter_rows(min_row=2, values_only=True):
        id_oferta, nombre_recurso = fila[1], fila[3]
        nombre = (nombre_recurso or "").strip()
        medico_por_oferta[id_oferta] = medicos_por_nombre.get(_normalizar_nombre(nombre)) if nombre else None

    creadas, omitidas = 0, 0
    nuevas = []
    for fila in ws_horarios.iter_rows(min_row=2, values_only=True):
        id_oferta = fila[0]
        hora_desde, hora_hasta = fila[2], fila[3]
        marcas_dias = fila[4:11]
        medico = medico_por_oferta.get(id_oferta)
        # Si el bloque no tiene servicios registrados, no se filtra por tipo (mejor
        # incluir de más que descartar de más por falta de dato).
        es_consulta = ofertas_consulta.get(id_oferta, True)
        if medico is None or not hora_desde or not hora_hasta or not es_consulta:
            omitidas += 1
            continue
        for indice_dia, marca in enumerate(marcas_dias):
            if marca:
                nuevas.append(OfertaMedico(
                    medico=medico,
                    dia_semana=indice_dia,
                    hora_inicio=_parsear_hora(hora_desde),
                    hora_fin=_parsear_hora(hora_hasta),
                ))
                creadas += 1

    with transaction.atomic():
        OfertaMedico.objects.all().delete()
        OfertaMedico.objects.bulk_create(nuevas)

    return creadas, omitidas


def importar_bloqueos_xlsx(contenido_bytes):
    wb = _cargar_libro(contenido_bytes)
    ws_bloqueos = wb["Bloqueos"]
    ws_horarios = wb["HorariosBloqueos"]
    medicos_por_nombre = _mapa_medicos_por_nombre()

    info_por_bloqueo = {}
    for fila in ws_bloqueos.iter_rows(min_row=2, values_only=True):
        id_bloqueo, nombre_recurso = fila[0], fila[2]
        tipo, motivo = fila[9], fila[10]
        nombre = (nombre_recurso or "").strip()
        info_por_bloqueo[id_bloqueo] = {
            "medico": medicos_por_nombre.get(_normalizar_nombre(nombre)) if nombre else None,
            "tipo": tipo or BloqueoMedico.Tipo.PARCIAL,
            "motivo": (motivo or "").strip(),
        }

    creados, omitidos = 0, 0
    nuevos = []
    for fila in ws_horarios.iter_rows(min_row=2, values_only=True):
        id_bloqueo = fila[0]
        hora_desde, hora_hasta = fila[2], fila[3]
        marcas_dias = fila[4:11]
        info = info_por_bloqueo.get(id_bloqueo)
        medico = info["medico"] if info else None
        if medico is None or not hora_desde or not hora_hasta:
            omitidos += 1
            continue
        for indice_dia, marca in enumerate(marcas_dias):
            if marca:
                nuevos.append(BloqueoMedico(
                    medico=medico,
                    dia_semana=indice_dia,
                    hora_inicio=_parsear_hora(hora_desde),
                    hora_fin=_parsear_hora(hora_hasta),
                    tipo=info["tipo"],
                    motivo=info["motivo"],
                ))
                creados += 1

    with transaction.atomic():
        BloqueoMedico.objects.all().delete()
        BloqueoMedico.objects.bulk_create(nuevos)

    return creados, omitidos
