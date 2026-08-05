"""Borra TODOS los médicos (y en cascada: sus asignaciones de box, citas
importadas, oferta y bloqueos) y los vuelve a cargar desde cero, desde la hoja
'LISTADO DE MEDICOS' de AGENDA.xlsx — la lista vigente del centro médico.

Es una operación destructiva e irreversible sobre datos ya cargados. Se usa
cuando la planilla fuente cambió lo suficiente (médicos nuevos, otros que ya
no corresponden) como para que reconciliar registro por registro no valga la
pena frente a partir de cero.

Uso: python manage.py reiniciar_medicos AGENDA.xlsx [--hoja "LISTADO DE MEDICOS"] [--dry-run]
"""

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from medicos.models import Especialidad, Medico


class Command(BaseCommand):
    help = "Borra todos los médicos y los recarga desde 'LISTADO DE MEDICOS' de AGENDA.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str)
        parser.add_argument("--hoja", type=str, default="LISTADO DE MEDICOS")
        parser.add_argument("--fila-inicio", type=int, default=2)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        try:
            wb = openpyxl.load_workbook(options["archivo"], data_only=True)
        except (OSError, KeyError) as exc:
            raise CommandError(f"No se pudo abrir '{options['archivo']}': {exc}")

        try:
            ws = wb[options["hoja"]]
        except KeyError:
            raise CommandError(
                f"La hoja '{options['hoja']}' no existe. Hojas disponibles: {wb.sheetnames}"
            )

        anteriores = Medico.objects.count()
        creados = 0
        omitidos = []

        with transaction.atomic():
            Medico.objects.all().delete()

            for fila in ws.iter_rows(min_row=options["fila_inicio"], values_only=True):
                nombre = (fila[0] or "").strip() if len(fila) > 0 else ""
                if not nombre:
                    continue
                rut = str(fila[1] or "").strip() if len(fila) > 1 else ""
                especialidad = (fila[2] or "").strip() if len(fila) > 2 else ""
                email = (fila[3] or "").strip() if len(fila) > 3 else ""
                telefono = str(fila[5] or "").strip() if len(fila) > 5 else ""
                tiempo = fila[8] if len(fila) > 8 else None

                if not rut:
                    omitidos.append((nombre, "sin RUT"))
                    continue

                try:
                    duracion = int(tiempo) if tiempo else 15
                except (TypeError, ValueError):
                    duracion = 15

                medico = Medico.objects.create(
                    nombre_completo=nombre,
                    rut=rut,
                    telefono=telefono,
                    email=email if "@" in email else "",
                    duracion_consulta_min=duracion,
                )
                if especialidad:
                    esp_obj, _ = Especialidad.objects.get_or_create(nombre=especialidad)
                    medico.especialidades.add(esp_obj)
                creados += 1

            if options["dry_run"]:
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f"Médicos anteriores eliminados: {anteriores}. Médicos nuevos creados: {creados}."
        ))
        if omitidos:
            self.stdout.write(self.style.WARNING("Filas omitidas (sin RUT):"))
            for nombre, motivo in omitidos:
                self.stdout.write(f"  - {nombre}: {motivo}")
        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-run: no se guardó nada (rollback)."))
