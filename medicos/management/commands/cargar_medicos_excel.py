"""Carga inicial de médicos desde la planilla real del centro (hoja 'Médicos CM' de
AGENDA.xlsx), incluyendo su box por defecto (regla fija) tal como venía en la columna BOX.

La columna BOX trae 3 formatos distintos, todos manejados aquí:
- Número (ej. 21): box fijo, aplica todos los días.
- Texto sin patrón de día (ej. "TORRE", "FUERA", "BOX ECOGINE"): se crea como Box con ese
  nombre, aplica todos los días.
- Texto con patrón "<días> <box>[, <días> <box>...]" (ej. "L-MA-MI-V 21, J 22"): reglas
  distintas por día de la semana.

Es idempotente: se puede re-ejecutar sin duplicar médicos, boxes ni asignaciones.

Uso: python manage.py cargar_medicos_excel AGENDA.xlsx
"""

import re

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from boxes.models import AsignacionBox, Box
from medicos.models import Especialidad, Medico

DIA_MAP = {
    "L": 0, "LU": 0,
    "MA": 1,
    "MI": 2, "X": 2,
    "J": 3, "JU": 3,
    "V": 4, "VI": 4,
    "S": 5, "SA": 5,
    "D": 6, "DO": 6,
}

PATRON_SEGMENTO = re.compile(r"^([A-ZÁÉÍÓÚ\-]+)\s+(\d+)$")


def _parsear_box(valor):
    """Devuelve una lista de tuplas (nombre_box, dia_semana|None), o None si es ambiguo."""
    if valor is None:
        return []
    if isinstance(valor, int):
        return [(f"Box {valor}", None)]
    if isinstance(valor, float):
        if valor.is_integer():
            return [(f"Box {int(valor)}", None)]
        return None  # ambiguo, ej. 16.29: requiere revisión manual

    texto = str(valor).strip()
    if not texto:
        return []

    resultado = []
    for segmento in (s.strip() for s in texto.split(",")):
        if not segmento:
            continue
        m = PATRON_SEGMENTO.match(segmento.upper())
        dias_codigos = m.group(1).split("-") if m else []
        if m and all(codigo in DIA_MAP for codigo in dias_codigos):
            box_nombre = f"Box {m.group(2)}"
            for codigo in dias_codigos:
                resultado.append((box_nombre, DIA_MAP[codigo]))
        else:
            resultado.append((segmento, None))
    return resultado


class Command(BaseCommand):
    help = "Carga médicos y sus boxes por defecto desde la hoja 'Médicos CM' de AGENDA.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str)
        parser.add_argument("--hoja", type=str, default="Médicos CM")
        parser.add_argument("--fila-inicio", type=int, default=4)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        try:
            wb = openpyxl.load_workbook(options["archivo"], data_only=True)
        except (OSError, KeyError) as exc:
            raise CommandError(f"No se pudo abrir '{options['archivo']}': {exc}")

        try:
            ws = wb[options["hoja"]]
        except KeyError:
            raise CommandError(
                f"La hoja '{options['hoja']}' no existe. Hojas disponibles: {wb.sheetnames}"
            )

        creados, actualizados, sin_box, ambiguos = 0, 0, 0, []

        with transaction.atomic():
            for fila in ws.iter_rows(min_row=options["fila_inicio"], values_only=True):
                nombre = (fila[1] or "").strip() if len(fila) > 1 else ""
                if not nombre:
                    continue
                rut = (fila[2] or "").strip() if len(fila) > 2 else ""
                especialidad = (fila[3] or "").strip() if len(fila) > 3 else ""
                email = (fila[4] or "").strip() if len(fila) > 4 else ""
                telefono = (fila[6] or "").strip() if len(fila) > 6 else ""
                box_raw = fila[7] if len(fila) > 7 else None

                if not rut:
                    self.stderr.write(self.style.WARNING(f"Omitido (sin RUT): {nombre}"))
                    continue

                medico, fue_creado = Medico.objects.update_or_create(
                    rut=rut,
                    defaults={
                        "nombre_completo": nombre,
                        "telefono": telefono,
                        "email": email if "@" in email else "",
                    },
                )
                creados += int(fue_creado)
                actualizados += int(not fue_creado)

                if especialidad:
                    esp_obj, _ = Especialidad.objects.get_or_create(nombre=especialidad)
                    medico.especialidades.add(esp_obj)

                reglas = _parsear_box(box_raw)
                if reglas is None:
                    ambiguos.append((nombre, rut, box_raw))
                    continue
                if not reglas:
                    sin_box += 1
                    continue

                for box_nombre, dia_semana in reglas:
                    box_obj, _ = Box.objects.get_or_create(nombre=box_nombre)
                    AsignacionBox.objects.get_or_create(
                        medico=medico, box=box_obj, dia_semana=dia_semana
                    )

            if options["dry_run"]:
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f"Médicos creados: {creados}, actualizados: {actualizados}, "
            f"sin box asignado: {sin_box}."
        ))
        if ambiguos:
            self.stdout.write(self.style.WARNING(
                "Filas con valor de BOX ambiguo (revisar y asignar a mano en el admin):"
            ))
            for nombre, rut, valor in ambiguos:
                self.stdout.write(f"  - {nombre} (RUT {rut}): BOX = {valor!r}")
        if options["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-run: no se guardó nada en la base."))
