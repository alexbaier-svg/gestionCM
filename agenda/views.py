import datetime
import re
import unicodedata

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

from disponibilidad.models import BloqueoMedico, OfertaMedico

from .forms import ImportarAgendaForm
from .importador import decodificar_bytes, importar_agenda_csv
from .models import BloqueAgenda, ImportacionAgenda

ESTADOS_VALIDOS = [
    BloqueAgenda.Estado.CITADO, BloqueAgenda.Estado.CONFIRMADO, BloqueAgenda.Estado.ATENDIDO,
]

# Especialidades a considerar en la agenda/resumen (definidas por el usuario).
_ESPECIALIDADES_VALIDAS_RAW = [
    "Cirugía Bariátrica", "Cirugía de Mama", "Cirugía General", "Coloproctología",
    "Gastroenterología Adulto", "Ginecología y Obstetricia", "Matrona",
    "Medicina Familiar", "Medicina General", "Medicina Interna", "Nefrología Adulto",
    "Neurología Infantil", "Nutricionista", "Otorrinolaringología", "Pediatría",
    "Psicología Adulto", "Psicología Infantil y Adolescente", "Traumatología - Hombro",
    "Traumatología Adulto", "Urología Adulto", "Neurología Adulto",
    "Matrona - Monitoreo Fetal", "Ginecología - Medicina Reproductiva e Infertilidad",
    "Cardiología Adulto", "Traumatología - Rodilla", "Cardiología Infantil",
    "Cirugía de Cabeza y Cuello", "Traumatología - Mano",
    "Medico con formación en Oftalmología", "Médico con Formación en Gastroenterología",
    "Traumatología - Codo", "Traumatología - Cadera", "Ginecología", "SAP Cadera MLE",
    "Mastología Los Andes",
]


def _normalizar_texto(texto):
    texto = unicodedata.normalize("NFKD", (texto or "").strip().lower())
    return texto.encode("ascii", "ignore").decode("ascii")


ESPECIALIDADES_VALIDAS = {_normalizar_texto(e) for e in _ESPECIALIDADES_VALIDAS_RAW}


def _box_numero(nombre):
    """'Box 10' -> '10'. Boxes con nombre no numérico (TORRE, BOX ECOGINE) quedan igual."""
    if not nombre:
        return "Sin asignar"
    m = re.match(r"^box\s+(\d+)$", nombre.strip(), re.IGNORECASE)
    return m.group(1) if m else nombre


def _bloques_del_dia(request, fecha):
    qs = (
        BloqueAgenda.objects.filter(fecha=fecha, estado__in=ESTADOS_VALIDOS)
        .select_related("medico", "box")
        .order_by("hora_inicio", "medico__nombre_completo")
    )
    medico_propio = getattr(request.user, "medico", None)
    es_gestion = request.user.groups.filter(name__in=["Administrador", "Recepción"]).exists()
    if medico_propio is not None and not request.user.is_superuser and not es_gestion:
        qs = qs.filter(medico=medico_propio)
    return [b for b in qs if _normalizar_texto(b.especialidad) in ESPECIALIDADES_VALIDAS]


def _minutos_libres(ofertas, bloqueos):
    """Minutos netos disponibles de un médico un día: sus ventanas de oferta,
    descontando lo que cubren sus bloqueos ese mismo día."""
    if any(b.tipo == BloqueoMedico.Tipo.DIA_COMPLETO for b in bloqueos):
        return 0

    total = 0
    for oferta in ofertas:
        segmentos = [(oferta.hora_inicio.hour * 60 + oferta.hora_inicio.minute,
                      oferta.hora_fin.hour * 60 + oferta.hora_fin.minute)]
        for b in bloqueos:
            bi = b.hora_inicio.hour * 60 + b.hora_inicio.minute
            bf = b.hora_fin.hour * 60 + b.hora_fin.minute
            nuevos = []
            for si, sf in segmentos:
                if bf <= si or bi >= sf:
                    nuevos.append((si, sf))
                    continue
                if bi > si:
                    nuevos.append((si, bi))
                if bf < sf:
                    nuevos.append((bf, sf))
            segmentos = nuevos
        total += sum(sf - si for si, sf in segmentos)
    return total


def _oferta_del_dia(dia_semana):
    """Ofertas y bloqueos de todos los médicos para un día de semana dado,
    agrupados por medico_id."""
    ofertas_por_medico = {}
    for o in OfertaMedico.objects.filter(dia_semana=dia_semana):
        ofertas_por_medico.setdefault(o.medico_id, []).append(o)
    bloqueos_por_medico = {}
    for b in BloqueoMedico.objects.filter(dia_semana=dia_semana):
        bloqueos_por_medico.setdefault(b.medico_id, []).append(b)
    return ofertas_por_medico, bloqueos_por_medico


def _resumen_por_medico(bloques, fecha):
    agrupado = {}
    for b in bloques:
        clave = (b.medico_id, b.box.nombre if b.box else None)
        if clave not in agrupado:
            agrupado[clave] = {
                "medico": b.medico, "hora_min": b.hora_inicio, "hora_max": b.hora_fin, "cantidad": 0,
            }
        info = agrupado[clave]
        info["hora_min"] = min(info["hora_min"], b.hora_inicio)
        info["hora_max"] = max(info["hora_max"], b.hora_fin)
        info["cantidad"] += 1

    ofertas_por_medico, bloqueos_por_medico = _oferta_del_dia(fecha.weekday())

    filas = []
    for (medico_id, box_nombre), info in agrupado.items():
        medico = info["medico"]
        minutos_libres = _minutos_libres(
            ofertas_por_medico.get(medico_id, []), bloqueos_por_medico.get(medico_id, [])
        )
        oferta_citas = minutos_libres // medico.duracion_consulta_min if medico.duracion_consulta_min else 0
        no_agendadas = max(oferta_citas - info["cantidad"], 0)
        filas.append({
            "medico": medico.nombre_completo,
            "box": _box_numero(box_nombre),
            "hora_min": info["hora_min"],
            "hora_max": info["hora_max"],
            "cantidad": info["cantidad"],
            "turno": "AM" if info["hora_min"] < datetime.time(13, 0) else "PM",
            "oferta_citas": oferta_citas,
            "no_agendadas": no_agendadas,
        })
    filas.sort(key=lambda f: f["medico"])
    return filas


@login_required
def agenda_diaria(request):
    fecha_str = request.GET.get("fecha")
    try:
        fecha = datetime.date.fromisoformat(fecha_str) if fecha_str else datetime.date.today()
    except ValueError:
        fecha = datetime.date.today()

    bloques = _bloques_del_dia(request, fecha)
    resumen = _resumen_por_medico(bloques, fecha)
    contexto = {
        "fecha": fecha,
        "fecha_anterior": fecha - datetime.timedelta(days=1),
        "fecha_siguiente": fecha + datetime.timedelta(days=1),
        "bloques": bloques,
        "resumen": resumen,
        "total_citas": sum(f["cantidad"] for f in resumen),
        "total_oferta": sum(f["oferta_citas"] for f in resumen),
        "total_no_agendadas": sum(f["no_agendadas"] for f in resumen),
    }
    return render(request, "agenda/diaria.html", contexto)


@login_required
def agenda_diaria_excel(request):
    import openpyxl
    from openpyxl.utils import get_column_letter

    fecha_str = request.GET.get("fecha")
    try:
        fecha = datetime.date.fromisoformat(fecha_str) if fecha_str else datetime.date.today()
    except ValueError:
        fecha = datetime.date.today()

    bloques = _bloques_del_dia(request, fecha)
    resumen = _resumen_por_medico(bloques, fecha)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen por médico"
    encabezados = [
        "Médico", "Box", "Turno", "Hora mín. desde", "Hora máx. hasta",
        "Cantidad de citas", "Oferta (citas)", "No agendadas",
    ]
    ws.append(encabezados)
    for fila in resumen:
        ws.append([
            fila["medico"],
            fila["box"],
            fila["turno"],
            fila["hora_min"].strftime("%H:%M"),
            fila["hora_max"].strftime("%H:%M"),
            fila["cantidad"],
            fila["oferta_citas"],
            fila["no_agendadas"],
        ])
    ws.append([
        "Total", "", "", "", "",
        sum(f["cantidad"] for f in resumen),
        sum(f["oferta_citas"] for f in resumen),
        sum(f["no_agendadas"] for f in resumen),
    ])
    for i, encabezado in enumerate(encabezados, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(16, len(encabezado) + 4)

    respuesta = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    respuesta["Content-Disposition"] = f'attachment; filename="agenda_{fecha.isoformat()}.xlsx"'
    wb.save(respuesta)
    return respuesta


@login_required
@permission_required("agenda.add_bloqueagenda", raise_exception=True)
def importar(request):
    if request.method == "POST":
        form = ImportarAgendaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            try:
                texto = decodificar_bytes(archivo.read())
            except UnicodeDecodeError:
                messages.error(
                    request,
                    "No se pudo leer el archivo. Debe ser el CSV exportado tal cual, sin "
                    "convertir su codificación.",
                )
                return render(request, "agenda/importar.html", {"form": form})

            importacion = importar_agenda_csv(texto, archivo.name, request.user)
            messages.success(
                request,
                f"Importación completa: {importacion.filas_procesadas} filas procesadas, "
                f"{importacion.filas_omitidas} omitidas (recursos/equipos sin médico "
                f"asociado, o citas ya importadas antes).",
            )
            primer_bloque = importacion.bloques.first()
            if primer_bloque:
                return redirect(f"/agenda/?fecha={primer_bloque.fecha.isoformat()}")
            return redirect("agenda:diaria")
    else:
        form = ImportarAgendaForm()

    importaciones_recientes = ImportacionAgenda.objects.order_by("-fecha_importacion")[:10]
    return render(request, "agenda/importar.html", {
        "form": form,
        "importaciones_recientes": importaciones_recientes,
    })
